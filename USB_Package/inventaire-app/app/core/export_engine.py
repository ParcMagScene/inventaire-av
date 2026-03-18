"""
export_engine.py — Export CSV et XLSX stylisé.
"""
import csv
import os
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from . import database as db
from .models import Article
from .totals_engine import TotalsEngine


class ExportEngine:
    """Exporte l'inventaire en CSV et XLSX."""

    def __init__(self, db_path=None):
        self.db_path = db_path
        self.totals = TotalsEngine(db_path)

    # ─── Export CSV ──────────────────────────────────
    def export_csv(self, output_path: str,
                   category_id: int | None = None,
                   location_id: int | None = None) -> str:
        """Exporte l'inventaire en CSV. Retourne le chemin du fichier."""
        articles = db.get_articles(self.db_path,
                                   category_id=category_id,
                                   location_id=location_id)

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        headers = [
            "Référence", "Nom", "Description", "Catégorie", "Emplacement",
            "Fournisseur", "Quantité", "Qté min",
            "Prix Bas", "Prix Moyen", "Prix Haut",
            "Total Bas", "Total Moyen", "Total Haut",
            "Mode prix", "Confiance", "Score confiance",
            "Source prix", "Notes", "Créé le", "Mis à jour le",
        ]

        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(headers)
            for a in articles:
                writer.writerow([
                    a.reference, a.name, a.description,
                    a.category_name, a.location_name, a.supplier_name,
                    a.quantity, a.quantity_min,
                    f"{a.price_low:.2f}", f"{a.price_avg:.2f}", f"{a.price_high:.2f}",
                    f"{a.total_low:.2f}", f"{a.total_avg:.2f}", f"{a.total_high:.2f}",
                    a.price_mode, a.confidence, a.confidence_score,
                    a.price_source, a.notes, a.created_at, a.updated_at,
                ])

            # Ligne de totaux
            gt = self.totals.global_totals(articles)
            writer.writerow([])
            writer.writerow([
                "TOTAUX", "", "", "", "", "",
                gt["total_quantity"], "",
                "", "", "",
                f"{gt['total_low']:.2f}", f"{gt['total_avg']:.2f}", f"{gt['total_high']:.2f}",
                "", "", "", "", "", "",
            ])

        return output_path

    # ─── Export XLSX stylisé ─────────────────────────
    def export_xlsx(self, output_path: str,
                    category_id: int | None = None,
                    location_id: int | None = None) -> str:
        """Exporte l'inventaire en XLSX avec mise en forme professionnelle."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side, numbers,
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl est requis pour l'export XLSX. "
                "Installez-le avec : pip install openpyxl"
            )

        articles = db.get_articles(self.db_path,
                                   category_id=category_id,
                                   location_id=location_id)

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        wb = Workbook()

        # ─── Couleurs et styles ──────────────────────
        TURQUOISE = "00BFA5"
        DARK_BG = "2B2D31"
        HEADER_BG = "363940"
        WHITE = "FFFFFF"
        LIGHT_GRAY = "F5F5F5"
        DANGER = "FF5252"
        WARNING = "FFAB40"
        SUCCESS = "69F0AE"

        header_font = Font(name="Segoe UI", bold=True, color=WHITE, size=10)
        header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        total_font = Font(name="Segoe UI", bold=True, color=WHITE, size=10)
        total_fill = PatternFill(start_color=TURQUOISE, end_color=TURQUOISE, fill_type="solid")
        even_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        normal_font = Font(name="Segoe UI", size=10)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        money_fmt = '#,##0.00 €'

        # ═══════════════════════════════════════════════
        #  Feuille 1 : Articles
        # ═══════════════════════════════════════════════
        ws = wb.active
        ws.title = "Inventaire"

        # Titre
        ws.merge_cells("A1:U1")
        title_cell = ws["A1"]
        title_cell.value = f"Inventaire AV — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        title_cell.font = Font(name="Segoe UI", bold=True, size=14, color=TURQUOISE)
        title_cell.alignment = Alignment(horizontal="center")

        headers = [
            "Référence", "Nom", "Description", "Catégorie", "Emplacement",
            "Fournisseur", "Quantité", "Qté min",
            "Prix Bas", "Prix Moyen", "Prix Haut",
            "Total Bas", "Total Moyen", "Total Haut",
            "Mode prix", "Confiance", "Score",
            "Source prix", "Notes", "Créé le", "Mis à jour le",
        ]

        # En-têtes row 3
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Données
        money_cols = {9, 10, 11, 12, 13, 14}  # colonnes de prix (1-indexed)
        for r, a in enumerate(articles, 4):
            values = [
                a.reference, a.name, a.description,
                a.category_name, a.location_name, a.supplier_name,
                a.quantity, a.quantity_min,
                a.price_low, a.price_avg, a.price_high,
                a.total_low, a.total_avg, a.total_high,
                a.price_mode, a.confidence, a.confidence_score,
                a.price_source, a.notes, a.created_at, a.updated_at,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = normal_font
                cell.border = thin_border
                if col in money_cols:
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal="right")
                if (r - 4) % 2 == 0:
                    cell.fill = even_fill

                # Coloration confiance
                if col == 16:  # Confiance
                    if val == "fort":
                        cell.font = Font(name="Segoe UI", bold=True, color=SUCCESS, size=10)
                    elif val == "moyen":
                        cell.font = Font(name="Segoe UI", bold=True, color=WARNING, size=10)
                    elif val == "faible":
                        cell.font = Font(name="Segoe UI", bold=True, color=DANGER, size=10)

            # Highlight stock bas
            if a.is_low_stock:
                ws.cell(row=r, column=7).font = Font(name="Segoe UI", bold=True, color=DANGER, size=10)

        # Ligne de totaux
        total_row = len(articles) + 4
        gt = self.totals.global_totals(articles)
        ws.cell(row=total_row, column=1, value="TOTAUX").font = total_font
        ws.cell(row=total_row, column=1).fill = total_fill
        for col in range(2, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = thin_border

        ws.cell(row=total_row, column=7, value=gt["total_quantity"]).font = total_font
        ws.cell(row=total_row, column=7).fill = total_fill
        for col_idx, key in [(12, "total_low"), (13, "total_avg"), (14, "total_high")]:
            cell = ws.cell(row=total_row, column=col_idx, value=gt[key])
            cell.font = total_font
            cell.fill = total_fill
            cell.number_format = money_fmt

        # Largeurs de colonnes
        col_widths = [12, 25, 20, 15, 15, 15, 8, 8, 12, 12, 12, 12, 12, 12, 10, 10, 6, 30, 15, 18, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Filtre auto
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{total_row - 1}"

        # ═══════════════════════════════════════════════
        #  Feuille 2 : Totaux par catégorie
        # ═══════════════════════════════════════════════
        ws2 = wb.create_sheet("Par catégorie")
        cat_totals = self.totals.totals_by_category(articles)
        self._write_summary_sheet(ws2, cat_totals, "Catégorie",
                                  header_font, header_fill, total_font, total_fill,
                                  normal_font, even_fill, thin_border, money_fmt)

        # ═══════════════════════════════════════════════
        #  Feuille 3 : Totaux par emplacement
        # ═══════════════════════════════════════════════
        ws3 = wb.create_sheet("Par emplacement")
        loc_totals = self.totals.totals_by_location(articles)
        self._write_summary_sheet(ws3, loc_totals, "Emplacement",
                                  header_font, header_fill, total_font, total_fill,
                                  normal_font, even_fill, thin_border, money_fmt)

        # ═══════════════════════════════════════════════
        #  Feuille 4 : Totaux par fournisseur
        # ═══════════════════════════════════════════════
        ws4 = wb.create_sheet("Par fournisseur")
        sup_totals = self.totals.totals_by_supplier(articles)
        self._write_summary_sheet(ws4, sup_totals, "Fournisseur",
                                  header_font, header_fill, total_font, total_fill,
                                  normal_font, even_fill, thin_border, money_fmt)

        wb.save(output_path)
        return output_path

    def _write_summary_sheet(self, ws, totals: dict, label: str,
                             header_font, header_fill, total_font, total_fill,
                             normal_font, even_fill, thin_border, money_fmt):
        """Écrit un onglet de totaux groupés."""
        from openpyxl.utils import get_column_letter

        headers = [label, "Quantité", "Valeur Basse", "Valeur Moyenne", "Valeur Haute"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for r, (name, vals) in enumerate(sorted(totals.items()), 2):
            ws.cell(row=r, column=1, value=name).font = normal_font
            ws.cell(row=r, column=2, value=vals["qty"]).font = normal_font
            for col, key in [(3, "low"), (4, "avg"), (5, "high")]:
                cell = ws.cell(row=r, column=col, value=vals[key])
                cell.font = normal_font
                cell.number_format = money_fmt
            if (r - 2) % 2 == 0:
                for c in range(1, 6):
                    ws.cell(row=r, column=c).fill = even_fill
            for c in range(1, 6):
                ws.cell(row=r, column=c).border = thin_border

        # Ligne de total
        tr = len(totals) + 2
        ws.cell(row=tr, column=1, value="TOTAL").font = total_font
        ws.cell(row=tr, column=1).fill = total_fill
        ws.cell(row=tr, column=2, value=sum(v["qty"] for v in totals.values())).font = total_font
        ws.cell(row=tr, column=2).fill = total_fill
        for col, key in [(3, "low"), (4, "avg"), (5, "high")]:
            cell = ws.cell(row=tr, column=col, value=sum(v[key] for v in totals.values()))
            cell.font = total_font
            cell.fill = total_fill
            cell.number_format = money_fmt
            cell.border = thin_border

        # Largeurs
        widths = [25, 10, 15, 15, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
